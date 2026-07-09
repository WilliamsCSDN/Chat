from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
from pymilvus import Collection, CollectionSchema, DataType, FieldSchema, connections, utility

import load_embedding


DEFAULT_EXCEL_PATH = Path(__file__).resolve().parent / "app_faq_1000.xlsx"
DEFAULT_COLLECTION = "app_faq"
DEFAULT_DIM = 384
DEFAULT_BATCH_SIZE = 200


@dataclass
class FaqRecord:
    faq_id: int
    category_l1: str
    category_l2: str
    category_l3: str
    question: str
    answer: str
    question_desc: str
    similar_questions: str
    searchable_text: str


def normalize_text(value, max_len: int) -> str:
    text = "" if value is None else str(value).strip()
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text[:max_len]


def parse_excel(path: Path) -> List[FaqRecord]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise ValueError(f"Excel 为空: {path}")

    header_map: Dict[str, int] = {str(name).strip(): idx for idx, name in enumerate(headers) if name is not None}
    required_headers = ["1级分类", "2级分类", "3级分类", "问题", "答案", "问题描述", "相似问"]
    missing = [h for h in required_headers if h not in header_map]
    if missing:
        raise ValueError(f"Excel 缺少必要列: {missing}")

    records: List[FaqRecord] = []
    for row_num, row in enumerate(rows, start=2):
        if not row or all((cell is None or str(cell).strip() == "") for cell in row):
            continue

        def value_of(name: str):
            idx = header_map[name]
            return row[idx] if idx < len(row) else None

        category_l1 = normalize_text(value_of("1级分类"), 128)
        category_l2 = normalize_text(value_of("2级分类"), 128)
        category_l3 = normalize_text(value_of("3级分类"), 128)
        question = normalize_text(value_of("问题"), 512)
        answer = normalize_text(value_of("答案"), 4096)
        question_desc = normalize_text(value_of("问题描述"), 1024)
        similar_questions = normalize_text(value_of("相似问"), 2048)
        similar_questions = " | ".join([part.strip() for part in similar_questions.split("\n") if part.strip()])

        searchable_parts = [
            f"分类:{' > '.join([c for c in [category_l1, category_l2, category_l3] if c])}",
            f"问题:{question}",
            f"问题描述:{question_desc}",
            f"相似问:{similar_questions}",
            f"答案:{answer}",
        ]
        searchable_text = "；".join(part for part in searchable_parts if part and not part.endswith(":"))
        searchable_text = normalize_text(searchable_text, 8192)

        records.append(
            FaqRecord(
                faq_id=row_num - 1,
                category_l1=category_l1,
                category_l2=category_l2,
                category_l3=category_l3,
                question=question,
                answer=answer,
                question_desc=question_desc,
                similar_questions=similar_questions,
                searchable_text=searchable_text,
            )
        )

    return records


def init_collection(collection_name: str, dim: int, drop_existing: bool) -> Collection:
    if utility.has_collection(collection_name):
        if drop_existing:
            Collection(collection_name).drop()
        else:
            return Collection(collection_name)

    fields = [
        FieldSchema("id", DataType.INT64, is_primary=True, auto_id=True),
        FieldSchema("faq_id", DataType.INT64),
        FieldSchema("category_l1", DataType.VARCHAR, max_length=128),
        FieldSchema("category_l2", DataType.VARCHAR, max_length=128),
        FieldSchema("category_l3", DataType.VARCHAR, max_length=128),
        FieldSchema("question", DataType.VARCHAR, max_length=512),
        FieldSchema("answer", DataType.VARCHAR, max_length=4096),
        FieldSchema("question_desc", DataType.VARCHAR, max_length=1024),
        FieldSchema("similar_questions", DataType.VARCHAR, max_length=2048),
        FieldSchema("searchable_text", DataType.VARCHAR, max_length=8192),
        FieldSchema("embedding", DataType.FLOAT_VECTOR, dim=dim),
    ]
    schema = CollectionSchema(fields, description="App FAQ knowledge base")
    return Collection(collection_name, schema=schema)


def insert_records(collection: Collection, records: List[FaqRecord], batch_size: int) -> None:
    model = load_embedding.load_embedding_model()
    embeddings = model.encode(
        [record.searchable_text for record in records],
        normalize_embeddings=True,
        show_progress_bar=True,
    ).tolist()

    for i in range(0, len(records), batch_size):
        batch = records[i : i + batch_size]
        batch_embeddings = embeddings[i : i + batch_size]
        collection.insert(
            [
                [r.faq_id for r in batch],
                [r.category_l1 for r in batch],
                [r.category_l2 for r in batch],
                [r.category_l3 for r in batch],
                [r.question for r in batch],
                [r.answer for r in batch],
                [r.question_desc for r in batch],
                [r.similar_questions for r in batch],
                [r.searchable_text for r in batch],
                batch_embeddings,
            ]
        )
        print(f"已插入: {min(i + batch_size, len(records))}/{len(records)}")

    collection.flush()


def ensure_index(collection: Collection) -> None:
    collection.create_index(
        field_name="embedding",
        index_params={
            "metric_type": "COSINE",
            "index_type": "IVF_FLAT",
            "params": {"nlist": 128},
        },
    )
    collection.load()


def main():
    parser = argparse.ArgumentParser(description="导入 FAQ Excel 到 Milvus")
    parser.add_argument("--excel-path", type=Path, default=DEFAULT_EXCEL_PATH, help="Excel 文件路径")
    parser.add_argument("--collection", default=DEFAULT_COLLECTION, help="Milvus collection 名")
    parser.add_argument("--host", default="localhost", help="Milvus host")
    parser.add_argument("--port", default="19530", help="Milvus port")
    parser.add_argument("--drop-existing", action="store_true", help="导入前先删除同名 collection")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE, help="批量写入大小")
    args = parser.parse_args()

    if not args.excel_path.exists():
        raise FileNotFoundError(f"Excel 不存在: {args.excel_path}")

    print(f"读取 Excel: {args.excel_path}")
    records = parse_excel(args.excel_path)
    if not records:
        raise ValueError("Excel 无可导入数据")
    print(f"加载 FAQ 条数: {len(records)}")

    print(f"连接 Milvus: {args.host}:{args.port}")
    connections.connect(alias="default", host=args.host, port=args.port)

    collection = init_collection(
        collection_name=args.collection,
        dim=DEFAULT_DIM,
        drop_existing=args.drop_existing,
    )
    print(f"Collection 就绪: {args.collection}")

    insert_records(collection, records, batch_size=args.batch_size)
    ensure_index(collection)
    print(f"导入完成，当前实体数: {collection.num_entities}")


if __name__ == "__main__":
    main()
